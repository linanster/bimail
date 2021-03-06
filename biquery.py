#! /usr/bin/env python3
# coding:utf8
#
import os
import sys
import openpyxl
import pymysql
import datetime
import decimal
import prettytable as pt

db_addr = 'estar.xlink.cloud'
db_port = 3306
db_user = 'root'
db_passwd = 'GEMysql20200'
db_name = 'GE'

FILENAME_EXCEL = './data/data_excel.xlsx'
FILENAME_TXT_PREVIEW = './data/overview.txt'
FILENAME_TXT_TABLE_DEVICE = './data/device.txt'
FILENAME_TXT_TABLE_DEVICETYPEFORMAT = './data/device_type_format.txt'
FILENAME_TXT_TABLE_CBYGE_DEVICES_COMMISSIONED = './data/cbyge_devices_commissioned.txt'

def create_db_conn():
    try:
        conn = pymysql.Connect(host=db_addr, port=db_port, user=db_user, passwd=db_passwd, db=db_name)
    except Exception as e:
        print(e)
        return -1
    else:
        print('==connect success==')
        return conn
def close_db_conn(conn):
    conn.close()
    print('==connect closed==')

def get_table_heads(desc):
    heads = list()
    for item in desc:
        heads.append(item[0])
    return heads

def write_sheet(sheet, heads, datass, row_start=1):
    # 1. write heads row
    row = row_start
    for col, item in enumerate(heads):
        sheet.cell(row, col+1).value = item
    # 2. write data rows
    for datas in datass:
        row += 1
        for col, item in enumerate(datas):
            if type(item) is datetime.date:
                item = item.strftime('%Y-%m-%d')
            if type(item) is datetime.datetime:
                item = item.strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(item, decimal.Decimal):
                item = int(item)
            print(item)
            sheet.cell(row, col+1).value = item
        
# sql query patterns
# sql_query_table_summary = 'select * from GE.summary order by statistics_time desc limit 5;'
sql_query_table_summary = 'select ID,STATISTICS_TIME,TOTAL_USER,TOTAL_AVS_USER,TOTAL_GVS_USER,TOTAL_WWA_DEVICE,TOTAL_WWG_DEVICE,TOTAL_ACTIVATED,TOTAL_UN_ACTIVATED from GE.summary order by statistics_time desc limit 5;'
sql_query_table_active = 'select * from GE.active order by statistics_time desc limit 5;'
sql_query_table_deviceproduced = 'select date(created_time) as produced_date,sum(produced_count) as produced_count from GE.device_produced group by produced_date order by produced_date desc limit 5;'
sql_query_table_device_total = 'select count(appliance_id) as "total" from GE.device;'
sql_query_table_device_agg_by_type = 'select device_type, count(appliance_id) as "count" from GE.device group by device_type order by count desc;'
sql_query_table_devicetypeformat = 'select * from GE.device_type_format;'
# add 12/28
sql_query_table_cbyge_devices_commisioned_agg = 'select statistics_time,sum(count) as devices_commissioned from GE.cbyge_devices_commissioned group by statistics_time order by statistics_time desc limit 5;'
sql_query_table_cbyge_devices_commissioned_latest = 'select * from cbyge_devices_commissioned where statistics_time=(select statistics_time from cbyge_devices_commissioned order by statistics_time desc limit 1);'

heads_summary = None
datas_summary = None
heads_active = None
datas_active = None
heads_deviceproduced = None
datas_deviceproduced = None
heads_device_total = None
datas_device_total = None
heads_device_agg_by_type = None
datas_device_agg_by_type = None
heads_devicetypeformat = None
datas_devicetypeformat = None
heads_cbyge_devices_commissioned_agg_sum = None
datas_cbyge_devices_commissioned_agg_sum = None
heads_cbyge_devices_commissioned_latest = None
datas_cbyge_devices_commissioned_latest = None


def query_db_data(conn):
    global heads_summary
    global datas_summary
    global heads_active
    global datas_active
    global heads_deviceproduced
    global datas_deviceproduced
    global heads_device_total
    global datas_device_total
    global heads_device_agg_by_type
    global datas_device_agg_by_type
    global heads_devicetypeformat
    global datas_devicetypeformat
    global heads_cbyge_devices_commissioned_agg_sum
    global datas_cbyge_devices_commissioned_agg_sum
    global heads_cbyge_devices_commissioned_latest
    global datas_cbyge_devices_commissioned_latest
    
    cursor = conn.cursor()
    
    
    cursor.execute(sql_query_table_summary)
    heads_summary = get_table_heads(cursor.description)
    datas_summary =  cursor.fetchall()
    # print(heads_summary)
    # print(datas_summary)
    
    cursor.execute(sql_query_table_active)
    heads_active = get_table_heads(cursor.description)
    datas_active =  cursor.fetchall()
    # print(heads_active)
    # print(datas_active)
    
    cursor.execute(sql_query_table_deviceproduced)
    heads_deviceproduced = get_table_heads(cursor.description)
    datas_deviceproduced =  cursor.fetchall()
    # print(heads_deviceproduced)
    # print(datas_deviceproduced)
    
    cursor.execute(sql_query_table_device_total)
    heads_device_total = get_table_heads(cursor.description)
    datas_device_total =  cursor.fetchall()
    # print(heads_device_total)
    # print(datas_device_total)
    
    cursor.execute(sql_query_table_device_agg_by_type)
    heads_device_agg_by_type = get_table_heads(cursor.description)
    datas_device_agg_by_type =  cursor.fetchall()
    # print(heads_device_agg_by_type)
    # print(datas_device_agg_by_type)
    
    cursor.execute(sql_query_table_devicetypeformat)
    heads_devicetypeformat = get_table_heads(cursor.description)
    datas_devicetypeformat =  cursor.fetchall()
    # print(heads_devicetypeformat)
    # print(datas_devicetypeformat)

    cursor.execute(sql_query_table_cbyge_devices_commisioned_agg)
    heads_cbyge_devices_commissioned_agg_sum = get_table_heads(cursor.description)
    datas_cbyge_devices_commissioned_agg_sum =  cursor.fetchall()
    # print(heads_cbyge_devices_commissioned_agg_sum)
    # print(datas_cbyge_devices_commissioned_agg_sum)

    cursor.execute(sql_query_table_cbyge_devices_commissioned_latest)
    heads_cbyge_devices_commissioned_latest = get_table_heads(cursor.description)
    datas_cbyge_devices_commissioned_latest =  cursor.fetchall()

    print('==query db data==')


# write into Excel
def write_data_excel():
    book = openpyxl.Workbook()
    sheet1 = book.create_sheet(index=0, title='Summary')
    sheet2 = book.create_sheet(index=1, title='Active')
    sheet3 = book.create_sheet(index=2, title='Device Produced')
    sheet4 = book.create_sheet(index=3, title='Device')
    
    write_sheet(sheet1, heads_summary, datas_summary)
    write_sheet(sheet2, heads_active, datas_active)
    write_sheet(sheet3, heads_deviceproduced, datas_deviceproduced)
    write_sheet(sheet4, heads_device_total, datas_device_total)
    write_sheet(sheet4, heads_device_agg_by_type, datas_device_agg_by_type, row_start=4)
    
    book.save(FILENAME_EXCEL)
    print('==write into excel==')

# write into txt as Table
def write_data_txt():
    tb_summary = pt.PrettyTable(field_names=heads_summary)
    tb_active = pt.PrettyTable(field_names=heads_active)
    tb_deviceproduced = pt.PrettyTable(field_names=heads_deviceproduced)
    tb_device_total = pt.PrettyTable(field_names=heads_device_total)
    tb_device_agg_by_type = pt.PrettyTable(field_names=heads_device_agg_by_type)
    tb_devicetypeformat = pt.PrettyTable(field_names=heads_devicetypeformat)
    tb_devicecommissioned_agg = pt.PrettyTable(field_names=heads_cbyge_devices_commissioned_agg_sum)
    tb_devicecommissioned_latest = pt.PrettyTable(field_names=heads_cbyge_devices_commissioned_latest)
    
    for data in datas_summary:
        tb_summary.add_row(data)
    
    for data in datas_active:
        tb_active.add_row(data)
    
    for data in datas_deviceproduced:
        tb_deviceproduced.add_row(data)
    
    for data in datas_device_total:
        tb_device_total.add_row(data)
    
    for data in datas_device_agg_by_type:
        tb_device_agg_by_type.add_row(data)
    
    for data in datas_devicetypeformat:
        tb_devicetypeformat.add_row(data)

    for data in datas_cbyge_devices_commissioned_agg_sum:
        tb_devicecommissioned_agg.add_row(data)

    for data in datas_cbyge_devices_commissioned_latest:
        tb_devicecommissioned_latest.add_row(data)
    # print(tb_summary)
    # print(tb_active)
    # print(tb_deviceproduced)
    # print(tb_device_total)
    # print(tb_device_agg_by_type)
    # print(tb_devicetypeformat)
    # print(tb_devicecommissioned)
    
    with open(FILENAME_TXT_PREVIEW, 'w') as file:
        file.writelines('== 1. Summary Table (daily update) ==\n')
        file.write(tb_summary.get_string())
        file.writelines('\n\n')
        file.writelines('== 2. Active Table (daily update) ==\n')
        file.write(tb_active.get_string())
        file.writelines('\n\n')
        file.writelines('== 3. Device_Produced Table (daily update) ==\n')
        file.write(tb_deviceproduced.get_string())
        file.writelines('\n\n')
        file.writelines('== 4. Cbyge_Devices_Commissioned Table (weekly update) ==\n')
        file.write(tb_devicecommissioned_agg.get_string())
        file.writelines('\n\n')
        file.writelines('== 5. Device Table ==\n')
        file.write(tb_device_total.get_string())

    with open(FILENAME_TXT_TABLE_DEVICE, 'w') as file:
        file.writelines('== Device Table (aggregated by type) ==\n')
        file.write(tb_device_agg_by_type.get_string())
        file.writelines('\n')

    with open(FILENAME_TXT_TABLE_DEVICETYPEFORMAT, 'w') as file:
        file.writelines('== Device_Type_Format Table ==\n')
        file.write(tb_devicetypeformat.get_string())
        file.writelines('\n')

    with open(FILENAME_TXT_TABLE_CBYGE_DEVICES_COMMISSIONED, 'w') as file:
        file.writelines('== Device_Type_Format Table (aggregated by type) ==\n')
        file.writelines('== This table is updated weekly by design ==\n')
        file.write(tb_devicecommissioned_latest.get_string())
        file.writelines('\n')

    print('==write into txt==')

def query_data():
    conn = create_db_conn()
    query_db_data(conn)
    close_db_conn(conn)

try:
    query_data()
    # write_data_excel()
    write_data_txt()
except Exception as e:
    print(e)
    sys.exit(1)

